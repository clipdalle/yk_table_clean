"""
Excel 列着色模块
根据配置字典为指定列着色
"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def color_excel_columns(excel_path: str, column_colors: dict, output_path: str = None):
    """
    根据配置字典为 Excel 文件的指定列着色

    Args:
        excel_path: 输入的 Excel 文件路径
        column_colors: 列着色配置字典，格式如 {'列名': '颜色代码', ...}
        output_path: 输出文件路径，如果为 None 则覆盖原文件
    """
    # 如果没有指定输出路径，则覆盖原文件
    if output_path is None:
        output_path = excel_path

    # 加载工作簿
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 定义颜色填充字典
    color_fills = {}
    for col_name, color_code in column_colors.items():
        # 跳过无颜色的列
        if not color_code or color_code == 'default':
            continue

        color_fills[col_name] = PatternFill(
            start_color=color_code,
            end_color=color_code,
            fill_type='solid'
        )

    # 获取表头行
    headers = [cell.value for cell in ws[1]]

    # 找到要着色的列索引
    columns_to_color = {}
    for col_name in column_colors.keys():
        if col_name in headers:
            col_index = headers.index(col_name) + 1  # 列索引从1开始
            columns_to_color[col_name] = (col_index, color_fills[col_name])

    # 应用着色
    colored_columns = []
    for col_name, (col_index, fill) in columns_to_color.items():
        col_letter = get_column_letter(col_index)

        # 对该列的所有单元格应用颜色（包括表头）
        for row in range(1, ws.max_row + 1):
            ws[f'{col_letter}{row}'].fill = fill

        colored_columns.append(f"{col_name}({col_letter}列)")

    # 保存文件
    wb.save(output_path)

    print(f"✅ 已为 {excel_path} 着色")
    print(f"   保存到: {output_path}")
    print(f"   着色列: {', '.join(colored_columns)}")
    print(f"   处理行数: {ws.max_row}")
    print(f"   未找到的列: {[col for col in column_colors.keys() if col not in headers]}")


def apply_column_colors_from_config(excel_path: str, config_dict: dict, output_path: str = None):
    """
    根据配置字典中的列颜色配置为Excel文件着色

    Args:
        excel_path: 输入的 Excel 文件路径
        config_dict: 配置字典，包含列颜色映射
        output_path: 输出文件路径，如果为 None 则覆盖原文件
    """
    # 从配置字典中提取列颜色映射
    column_colors = {}
    for item in config_dict:
        col_name = item.get('col_name')
        col_color = item.get('col_color')
        if col_name and col_color and col_color not in ('default', None):
            column_colors[col_name] = col_color

    # 调用着色函数
    color_excel_columns(excel_path, column_colors, output_path)


def apply_column_colors_from_config_memory(ws, config_dict: dict):
    """
    根据配置字典中的列颜色配置为工作表着色（内存版本）
    
    Args:
        ws: openpyxl工作表对象
        config_dict: 配置字典，包含列颜色映射
    """
    from openpyxl.styles import PatternFill
    
    # 从配置字典中提取列颜色映射
    column_colors = {}
    for item in config_dict:
        col_name = item.get('col_name')
        col_color = item.get('col_color')
        if col_name and col_color and col_color not in ('default', None):
            column_colors[col_name] = col_color
    
    # 获取表头行
    headers = [cell.value for cell in ws[1]]
    
    # 应用着色
    for col_name, color_code in column_colors.items():
        if col_name in headers:
            col_index = headers.index(col_name) + 1
            col_letter = get_column_letter(col_index)
            
            fill = PatternFill(
                start_color=color_code,
                end_color=color_code,
                fill_type='solid'
            )
            
            # 对该列的所有单元格应用颜色
            for row in range(1, ws.max_row + 1):
                ws[f'{col_letter}{row}'].fill = fill


def apply_color_by_value_memory(ws, value_color_mapping, match_mode='exact'):
    """
    根据值着色单元格（内存版本）
    
    Args:
        ws: openpyxl工作表对象
        value_color_mapping: 值颜色映射列表
        match_mode: 匹配模式 'exact' 或 'contains_value'
    """
    from openpyxl.styles import PatternFill
    
    # 创建颜色填充对象
    color_fills = {}
    for row in value_color_mapping:
        color_fills[row['cell_value']] = PatternFill(
            start_color=row['color_code'],
            end_color=row['color_code'],
            fill_type='solid'
        )
    
    # 遍历所有单元格
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_value_str = str(cell.value)
                
                # 根据匹配模式进行匹配
                for value, fill in color_fills.items():
                    if match_mode == 'exact' and cell_value_str == value:
                        cell.fill = fill
                    elif match_mode == 'contains_value' and value in cell_value_str:
                        cell.fill = fill
