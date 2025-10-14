"""
排麦人员 & 主持人员字段清洗 Pipeline
使用 Gemini API 批量解析混乱的文本，生成人工校验表
"""

import pandas as pd
import json
import time
import sys
import os
import tempfile
import traceback
import re
import io
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
from tenacity import retry, stop_after_attempt, wait_fixed
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pipeline.llm.llm_client import _get_client
from pipeline.prompts_v3 import PROMPT_BATCH
from pipeline.utils import (normalize_ascii_lower, dedup_names, extract_chinese, 
                             further_split, parse_json_markdown)
from global_config import TEMPERATURE, BATCH_SIZE, CURRENT_MODEL, COLS_CONFIG, STRICT_DATE_FILTER 


# Excel 配置常量
DATE_FROM_UI_CELL = 'V1'  # 用户指定日期的单元格位置
DB_NAME_LIST_DIR = 'name_list'

def get_temp_path(filename):
    temp_dir = tempfile.gettempdir()
    return os.path.join(temp_dir, filename)
    


model_client = _get_client(model_name=CURRENT_MODEL)




@retry(stop=stop_after_attempt(3), wait=wait_fixed(5))
def _parse_single_batch(batch_df: pd.DataFrame, assist_known_names: List[str]) -> List[Dict]:
    """
    内部函数：处理单个批次（一次 LLM 调用）
    
    Args:
        batch_df: 批次数据框
        assist_known_names: 辅助人名列表
    
    Returns:
        解析结果列表
    """
    # 构建批量输入数据
    if '行号' not in batch_df.columns:
        raise ValueError('行号列不存在')
    if '厅名中文' not in batch_df.columns:
        raise ValueError('厅名中文列不存在')
    if '日期（必填）' not in batch_df.columns:
        raise ValueError('日期（必填）列不存在')
    if '主持（必填）' not in batch_df.columns:
        raise ValueError('主持（必填）列不存在')
    if '排麦人员（必填）' not in batch_df.columns:
        raise ValueError('排麦人员（必填）列不存在')
    
    batch_data = []
    for idx, row in batch_df.iterrows():
        batch_data.append({
            "行号": idx,
            "厅名中文": row['厅名中文'],
            "日期": row['日期（必填）'],
            "主持": row['主持（必填）'],
            "排麦人员": row['排麦人员（必填）']
        })
    
    # 构建 prompt
    prompt = PROMPT_BATCH.format(
        batch_data=json.dumps(batch_data, ensure_ascii=False, indent=2),
        known_names=json.dumps(assist_known_names, ensure_ascii=False, indent=2)
    )
    
 
    try:
        response = model_client.get_completion(prompt, temperature=TEMPERATURE)
        
        # 清理 markdown 代码块
        response = response.strip()
        if response.startswith('```json'):
            response = response[7:]
        if response.startswith('```'):
            response = response[3:]
        if response.endswith('```'):
            response = response[:-3]
        response = response.strip()
        
        # 解析 JSON
        results = json.loads(response)
    except:
        print(response)
        traceback.print_exc()
        raise 
    
    # 验证结果
    if not isinstance(results, list):
        raise ValueError("返回结果不是数组")
    
    return results
     

def parse_hall_df(
    df: pd.DataFrame, 
    assist_known_names: List[str] = None) -> pd.DataFrame:
    """
    解析排麦人员和主持人员（自动分批，LLM处理，写回DataFrame）
    日期匹配和过滤通过后续步骤（公式+灰显）处理
    """
    total = len(df)
    num_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE

    print(f"开始批量解析 {total} 条记录...")
    print(f"批次设置: 每批 {BATCH_SIZE} 行，共 {num_batches} 批")
    print(f"预计耗时: 约 {num_batches * 0.5:.1f} 分钟")

    # 确保目标列存在
    target_columns = [
        '主持人员列表_AI解析',
        '排麦人员列表_AI解析',
        '排麦出席人数_AI解析',
        '排麦缺席人数_AI解析',
        '排麦置信度_AI解析',
        '标准化日期_AI解析',
        '数据日期和用户指定日期是否匹配',
    ]
    for col in target_columns:
        if col not in df.columns:
            df[col] = ''

    total_start_ts = time.time()

    for batch_idx in range(num_batches):
        start_idx = batch_idx * BATCH_SIZE
        end_idx = min((batch_idx + 1) * BATCH_SIZE, total)

        batch_df = df.iloc[start_idx:end_idx]

        print(f"处理批次 {batch_idx + 1}/{num_batches} (第 {start_idx+1}-{end_idx} 行)...", end='')
        batch_start_ts = time.time()

        batch_results = _parse_single_batch(batch_df, assist_known_names=assist_known_names)
            
        # 2) 逐行写回
        for result in batch_results:
            idx = result.get('行号')
            host_data = result['主持']
            paimai_data = result['排麦']
            error_msg = result.get('错误', '')

            host_list = further_split(host_data['主持人员列表'], split_char='-')
            paimai_list = further_split(paimai_data['排麦人员列表'], split_char='-')
            lack_num = paimai_data['缺席人数']
            conf = paimai_data['置信度']
            row_standard_date_str = result['标准化日期']

            df.at[idx, '主持人员列表_AI解析'] = '|'.join(host_list)
            df.at[idx, '排麦人员列表_AI解析'] = '|'.join(paimai_list)
            df.at[idx, '排麦出席人数_AI解析'] = len(paimai_list)
            df.at[idx, '排麦缺席人数_AI解析'] = lack_num
            df.at[idx, '排麦置信度_AI解析'] = conf
            df.at[idx, '标准化日期_AI解析'] = row_standard_date_str

        elapsed = time.time() - batch_start_ts
        print(f" 完成，用时 {elapsed:.2f}s")
        time.sleep(1)



        # 限流保护
        if batch_idx < num_batches - 1:
            time.sleep(0.5)

    total_elapsed = time.time() - total_start_ts
    print(f"\n✅ 解析完成！总用时 {total_elapsed/60:.2f} 分钟（{total_elapsed:.1f} 秒）")
    
    return df


def save_cleaned_data(df: pd.DataFrame, output_path: str):
    """
    保存清洗后的数据（按 COLS_CONFIG 排序列）

    Args:
        df: 包含解析结果的数据框
        output_path: 输出文件路径
    """
    # 按 COLS_CONFIG 重新排列列
    col_order = [col['col_name'] for col in COLS_CONFIG if col['col_name'] in df.columns]
    # 添加不在 COLS_CONFIG 中的列（如果有）
    remaining_cols = [col for col in df.columns if col not in col_order]
    final_col_order = col_order + remaining_cols

    df_sorted = df[final_col_order]

    # 保存为普通 Excel
    df_sorted.to_excel(output_path, index=False)

    print(f"✅ 清洗后的数据已保存: {output_path}")
    print(f"   - 原始列: {len(df_sorted.columns) - 7} 列")
    print(f"   - 新增列: 7 列（主持x1、排麦x4、错误x2）")
    print(f"   - 总列数: {len(df_sorted.columns)} 列")
    print(f"   - 列已按 CONFIG.COLS_CONFIG 排序")


def apply_gray_fill_for_date_mismatch(excel_path: str):
    """
    为日期不匹配的行添加灰色背景（A到S列）
    
    Args:
        excel_path: Excel文件路径
    """
    # data_only=True 会读取公式的计算结果而不是公式本身
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # 找到"数据日期和用户指定日期是否匹配"列和"标准化日期_AI解析"列
    headers = [cell.value for cell in ws[1]]
    match_col_idx = None
    standard_date_col_idx = None
    
    for idx, header in enumerate(headers, 1):
        if header == '数据日期和用户指定日期是否匹配':
            match_col_idx = idx
        elif header == '标准化日期_AI解析':
            standard_date_col_idx = idx
    
    if not match_col_idx or not standard_date_col_idx:
        wb.close()
        return
    
    # 获取用户指定日期
    user_date = ws[DATE_FROM_UI_CELL].value
    
    # 重新加载工作簿以便写入（data_only模式是只读的计算结果）
    wb.close()
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 创建灰显样式 - 现代灰显主题
    gray_fill = PatternFill(
        start_color='F0F0F0',  # 更浅的灰色背景
        end_color='F0F0F0',
        fill_type='solid'
    )
    
    # 创建灰色字体样式
    gray_font = Font(
        color='808080',  # 深灰色文字
        size=10          # 稍小的字体（可选，让其更"次要"）
    )
    
    # 遍历数据行（从第2行开始）
    colored_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        # 直接比较标准化日期和用户指定日期
        standard_date = ws.cell(row=row_idx, column=standard_date_col_idx).value
        if standard_date != user_date:
            # 为该行从A列到S列（第1-19列）添加灰显样式
            for col_idx in range(1, 20):  # S列是第19列
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = gray_fill    # 浅灰背景
                cell.font = gray_font    # 深灰文字
            colored_rows += 1
    
    wb.save(excel_path)
    wb.close()
    print(f"✅ 已为 {colored_rows} 行日期不匹配的数据添加灰色背景（A-S列）") 


def add_date_match_formulas(excel_path: str, date_str_from_ui: str):
    """
    在Z列添加日期匹配公式
    
    Args:
        excel_path: Excel文件路径
        date_str_from_ui: 用户指定的日期字符串
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    ws[DATE_FROM_UI_CELL] = date_str_from_ui
    
    # 找到标准化日期_AI解析列和数据日期和用户指定日期是否匹配列
    headers = [cell.value for cell in ws[1]]
    standard_date_col = None
    match_col = None
    
    for idx, header in enumerate(headers, 1):
        if header == '标准化日期_AI解析':
            standard_date_col = get_column_letter(idx)
        elif header == '数据日期和用户指定日期是否匹配':
            match_col = get_column_letter(idx)
    
    if standard_date_col and match_col:
        # 为每一行的"数据日期和用户指定日期是否匹配"列添加公式
        # 使用配置变量 DATE_FROM_UI_CELL 而不是写死的单元格引用
        # 提取列字母和行号，添加 $ 符号使行号绝对引用
        date_cell_col = ''.join([c for c in DATE_FROM_UI_CELL if c.isalpha()])
        date_cell_row = ''.join([c for c in DATE_FROM_UI_CELL if c.isdigit()])
        date_cell_ref = f'{date_cell_col}${date_cell_row}'  # 如 V$1
        
        for row_idx in range(2, ws.max_row + 1):
            formula = f'=IF({standard_date_col}{row_idx}={date_cell_ref},TRUE,FALSE)'
            ws[f'{match_col}{row_idx}'] = formula
    
    wb.save(excel_path)
    wb.close()
    print(f"✅ 已添加日期匹配公式（参考单元格: {DATE_FROM_UI_CELL}），用户指定日期: {date_str_from_ui}")


def apply_color_by_value(excel_path, value_color_mapping, output_path=None):
    """
    根据单元格值进行着色，每个配置项支持独立的匹配模式
    
    Args:
        excel_path: Excel文件路径
        value_color_mapping: 值-颜色映射列表，格式如：
            [
                {'cell_value': 'Siri', 'color_code': 'FF0000', 'mode': 'exact'},      # 精确匹配
                {'cell_value': '王摆摆', 'color_code': '00FF00', 'mode': 'contains_value'},    # 包含匹配
                {'cell_value': 'error', 'color_code': 'FFFF00', 'mode': 'exact'}      # 精确匹配
            ]
        output_path: 输出文件路径，如果为None则覆盖原文件
    """
    import openpyxl
    from openpyxl.styles import PatternFill
    
    # 如果没有指定输出路径，则覆盖原文件
    if output_path is None:
        output_path = excel_path
    
    # 加载工作簿
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # 遍历所有单元格进行着色
    colored_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_value_str = str(cell.value).strip()
                
                # 遍历每个配置项，使用其独立的匹配模式
                for config in value_color_mapping:
                    cell_value = config['cell_value']
                    color_code = config['color_code']
                    mode = config.get('mode', 'exact')  # 默认使用精确匹配
                    
                    # 根据匹配模式进行判断
                    should_color = False
                    if mode == 'exact':
                        # 精确匹配：str == str
                        should_color = (cell_value_str == cell_value)
                    elif mode == 'contains_value':
                        # 包含匹配：cell_value in cell_value_str
                        should_color = (cell_value in cell_value_str)
                    else:
                        raise ValueError(f"不支持的匹配模式: {mode}，请使用 'exact' 或 'contains_value'")
                    
                    # 如果匹配成功，进行着色
                    if should_color:
                        cell.fill = PatternFill(
                            start_color=color_code,
                            end_color=color_code,
                            fill_type='solid'
                        )
                        colored_count += 1
                        break  # 找到一个匹配就停止，避免重复着色
    
    # 保存文件
    wb.save(output_path)
    wb.close()
    
    print(f"✅ 已根据值着色完成")
    print(f"   - 输入文件: {excel_path}")
    print(f"   - 输出文件: {output_path}")
    print(f"   - 着色单元格数: {colored_count}")
    print(f"   - 颜色映射: {len(value_color_mapping)} 个")
    
    return colored_count



def save_cleaned_data_with_formula(df: pd.DataFrame, output_path: str, all_known_names: List[str], date_str_from_ui: str):
    """
    保存清洗后的数据（带公式版本，统计表在右侧）
    
    Args:
        df: 包含解析结果的数据框
        output_path: 输出文件路径
        all_known_names: 所有已知人名列表
        date_str_from_ui: 从UI传入的日期字符串
    """

    # 按 COLS_CONFIG 重新排列列
    col_order = [col['col_name'] for col in COLS_CONFIG if col['col_name'] in df.columns]
    remaining_cols = [col for col in df.columns if col not in col_order]
    final_col_order = col_order + remaining_cols
    df_sorted = df[final_col_order]

    # 保存第一步：清洗后的数据
    path_step1 = output_path.replace('.xlsx', '.step1_clean_table.xlsx')
    df_sorted.to_excel(path_step1, index=False)


    add_date_match_formulas(str(path_step1), date_str_from_ui)
 
    
    # 加载工作簿
    wb = load_workbook(path_step1)
    ws = wb.active

    # 找到相关列（用于右侧统计区）
    headers = [cell.value for cell in ws[1]]
    host_list_col = None
    paimai_list_col = None
    
    for idx, header in enumerate(headers, 1):
        if header == '主持人员列表_AI解析':
            host_list_col = idx
        elif header == '排麦人员列表_AI解析':
            paimai_list_col = idx
    
    # 初始化sorted_names为空列表
    sorted_names = []
    
    if host_list_col and paimai_list_col:
        host_col_letter = get_column_letter(host_list_col)
        paimai_col_letter = get_column_letter(paimai_list_col)
        data_row_count = ws.max_row
        
        # 提取所有唯一的人名（从主持和排麦列表中）
        all_names = set()
        
        # 从主持列提取（使用竖线分隔）
        for row in range(2, data_row_count + 1):
            cell_value = ws[f'{host_col_letter}{row}'].value
            if cell_value and str(cell_value).strip():
                names = [name.strip() for name in str(cell_value).split('|')]
                all_names.update(names)
        
        # 从排麦列提取（使用竖线分隔）
        for row in range(2, data_row_count + 1):
            cell_value = ws[f'{paimai_col_letter}{row}'].value
            if cell_value and str(cell_value).strip():
                names = [name.strip() for name in str(cell_value).split('|')]
                all_names.update(names)
        
        # 按字母顺序排序
        all_names = dedup_names(all_names)
        sorted_names = sorted(all_names)
 
        
        # 统计表放在右侧，空两列后开始
        stats_start_col = len(headers) + 3  # 空两列
        stats_col_1 = get_column_letter(stats_start_col)
        stats_col_2 = get_column_letter(stats_start_col + 1)
        stats_col_3 = get_column_letter(stats_start_col + 2)
        
        # 设置统计表表头（带样式）
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 深蓝色
        header_font = Font(bold=True, color='FFFFFF')  # 白色粗体
        
        ws[f'{stats_col_1}1'] = '人员姓名'
        ws[f'{stats_col_1}1'].fill = header_fill
        ws[f'{stats_col_1}1'].font = header_font
        ws[f'{stats_col_1}1'].alignment = Alignment(horizontal='center')
        
        ws[f'{stats_col_2}1'] = '主持次数'
        ws[f'{stats_col_2}1'].fill = header_fill
        ws[f'{stats_col_2}1'].font = header_font
        ws[f'{stats_col_2}1'].alignment = Alignment(horizontal='center')
        
        ws[f'{stats_col_3}1'] = '排麦次数'
        ws[f'{stats_col_3}1'].fill = header_fill
        ws[f'{stats_col_3}1'].font = header_font
        ws[f'{stats_col_3}1'].alignment = Alignment(horizontal='center')
        
        # 找到"数据日期和用户指定日期是否匹配"列
        match_col_letter = None
        for idx, header in enumerate(headers, 1):
            if header == '数据日期和用户指定日期是否匹配':
                match_col_letter = get_column_letter(idx)
                break
        
        # 填充人名和公式
        for idx, name in enumerate(sorted_names, 2):  # 从第2行开始
            # 人员姓名
            ws[f'{stats_col_1}{idx}'] = name
            
            if match_col_letter:
                # 主持次数（COUNTIFS 公式，只统计日期匹配的行）
                formula_host = f'=COUNTIFS({host_col_letter}:{host_col_letter},"*{name}*",{match_col_letter}:{match_col_letter},TRUE)'
                ws[f'{stats_col_2}{idx}'] = formula_host
                
                # 排麦次数（COUNTIFS 公式，只统计日期匹配的行）
                formula_paimai = f'=COUNTIFS({paimai_col_letter}:{paimai_col_letter},"*{name}*",{match_col_letter}:{match_col_letter},TRUE)'
                ws[f'{stats_col_3}{idx}'] = formula_paimai
            else:
                # 如果没有匹配列，使用旧的COUNTIF公式（兼容性）
                formula_host = f'=COUNTIF({host_col_letter}:{host_col_letter},"*{name}*")'
                ws[f'{stats_col_2}{idx}'] = formula_host
                
                formula_paimai = f'=COUNTIF({paimai_col_letter}:{paimai_col_letter},"*{name}*")'
                ws[f'{stats_col_3}{idx}'] = formula_paimai
        
        # 设置列宽
        ws.column_dimensions[stats_col_1].width = 20
        ws.column_dimensions[stats_col_2].width = 15
        ws.column_dimensions[stats_col_3].width = 15

    # 保存
    path_step2 = output_path.replace('.xlsx', '.add_stats_formula.xlsx')
    wb.save(path_step2)
    wb.close()

    # 直接对最终文件进行着色（不生成额外文件）
    path_step3 = output_path.replace('.xlsx', '.step3_nowarn.output.xlsx')
    from pipeline.coloring import apply_column_colors_from_config
    apply_column_colors_from_config(path_step2, COLS_CONFIG, path_step3)

    cell_config_list = [
        {'cell_value': 'medium', 'color_code': 'FF0000','mode': 'exact'},
        {'cell_value': 'low', 'color_code': 'FF0000','mode': 'exact'},
        {'cell_value': '9', 'color_code': 'FF0000','mode': 'exact'},
    ]
    for stat_name in sorted_names:
        if normalize_ascii_lower(stat_name) not in all_known_names:
            cell_config_list.append({'cell_value': stat_name, 'color_code': 'FF0000','mode': 'contains_value'})

    apply_color_by_value(path_step3, cell_config_list, output_path=output_path)
    
    # 为日期不匹配的行添加灰色背景
    apply_gray_fill_for_date_mismatch(output_path)

    print(f"✅ 已保存（带公式且着色）: {output_path}")
    print(f"   - 数据列: {len(headers)} 列")
    print(f"   - 统计表: 在右侧（空2列后）")
    print(f"   - 统计人数: {len(sorted_names) if host_list_col and paimai_list_col else 0} 人")
    print(f"   - 公式会自动统计每个人的主持次数和排麦次数")
    print(f"   - 修改人员列表后，统计会自动更新")
    print(f"   - 基于 CONFIG.COLS_CONFIG 配置着色")

    #清理step1和step2中间文件
    Path(path_step1).unlink()
    Path(path_step2).unlink()
    Path(path_step3).unlink()



def generate_statistics(df: pd.DataFrame):
    """
    生成统计报告
    
    Args:
        df: 包含解析结果的数据框
    """
    print("\n" + "="*60)
    print("解析结果统计")
    print("="*60)
    
    total = len(df)
    
    # 主持字段统计
    print("\n【主持字段】")
    host_errors = (df['主持错误_AI解析'] != '').sum()
    print(f"成功解析: {total - host_errors} ({(total-host_errors)/total*100:.1f}%)")
    print(f"解析错误: {host_errors}")
    
    # 排麦字段统计
    print("\n【排麦字段】")
    paimai_high = (df['排麦置信度_AI解析'] == 'high').sum()
    paimai_medium = (df['排麦置信度_AI解析'] == 'medium').sum()
    paimai_low = (df['排麦置信度_AI解析'] == 'low').sum()
    paimai_errors = (df['排麦错误_AI解析'] != '').sum()
    
    print(f"高置信度: {paimai_high} ({paimai_high/total*100:.1f}%)")
    print(f"中置信度: {paimai_medium} ({paimai_medium/total*100:.1f}%)")
    print(f"低置信度: {paimai_low} ({paimai_low/total*100:.1f}%)")
    print(f"解析错误: {paimai_errors}")
    
    # 建议
    need_check = paimai_low + paimai_medium
    print(f"\n建议优先检查: 约 {need_check} 条记录（排麦低+中置信度）")

# ========== 主流程 ========== 
def extract_chinese(text):
    return re.sub(r'[^\u4e00-\u9fa5]', '', text)



def get_hall_2_names_from_db():
    hall_2_names = {}
    for p in Path(DB_NAME_LIST_DIR).glob('*.txt'):
        with open(p, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f.read().splitlines() if line.strip()]
            hall_2_names[p.stem] = lines
    return hall_2_names
def gen_all_known_names(known_names_from_ui: List[str], selected_halls: List[str]):
    hall_2_names = get_hall_2_names_from_db()
    known_names_from_local = []
    for hall in selected_halls:
        known_names_from_local.extend(hall_2_names[hall])

    
    all_names = []
    if known_names_from_ui:
        all_names.extend(known_names_from_ui)
    all_names.extend(known_names_from_local)
    all_names = list(set(all_names))
    return all_names
 

def preprocess_df(df: pd.DataFrame):
    df['行号'] = df.index + 2
    df['厅名中文'] = df['厅号（必填）'].apply(extract_chinese)
    return df

def process_one_file(
        excel_path: str, 
        output_path: str, 
        date_str_from_ui: str, 
        strict_date_filter: bool = False, 
        selected_halls: list = None,
        known_names_from_ui: List[str] = None):
    """处理单个 Excel 文件并输出最终着色文件

    Args:
        excel_path: 输入 Excel 路径
        output_path: 最终着色文件路径（必填）。
        date_str_from_ui: 日期字符串
        strict_date_filter: 是否严格按日期筛选
        selected_halls: 选择的厅号列表，用于筛选数据
        known_names_from_ui: 从前端上传的已知人名列表
    """
    output_dir = Path(output_path).parent
    # 确保只在temp目录中创建目录
    import tempfile
    temp_dir = tempfile.gettempdir()
    if not str(output_dir).startswith(temp_dir):
        output_dir = Path(tempfile.mkdtemp())
        output_path = str(output_dir / Path(output_path).name)
    output_dir.mkdir(exist_ok=True)
    print("="*60)
    print("排麦人员 & 主持人员字段清洗 Pipeline")
    print("="*60)
    print(f"当前使用模型: {CURRENT_MODEL.upper()}")
    print("="*60)
    
    # 1. 读取数据
    print(f"\n📖 读取数据: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name=0)
    print(excel_path, '+++++++++++++++')
    df['厅名中文'] = df['厅号（必填）'].apply(extract_chinese)
    print(f"   - 原始记录数: {len(df)}")
        
    # 筛选特定厅号（必须选择厅号）
    if not selected_halls or len(selected_halls) == 0:
        raise ValueError("❌ 错误：必须至少选择一个厅号进行处理")
    
    # 从 name_list 目录读取辅助名单
    all_known_names = gen_all_known_names(known_names_from_ui, selected_halls)
        
    # 将厅号列表拼接成正则表达式，例如：['醉春色', '百媚生'] -> '醉春色|百媚生'
    hall_filter = '|'.join(selected_halls)
    print(f"   - 使用选择的厅号筛选: {hall_filter}")
    df = df[df['厅号（必填）'].str.contains(hall_filter, na=False)]
    print(f"✅ 读取完成，筛选后共 {len(df)} 条记录")

    # 2. 批量解析
    print(f"\n🤖 开始 LLM 批量解析...")
    df_parsed = parse_hall_df(
        df,
        assist_known_names=all_known_names
    )

    # 3. 生成统计
    generate_statistics(df_parsed)
    
    # 4. 在Z列添加日期匹配公式
 
    
    # 5. 仅保存带公式版本并生成着色版本（中间产物）
    print(f"\n📊 保存带公式版本并着色...")
    save_cleaned_data_with_formula(df_parsed, str(output_path), all_known_names, date_str_from_ui)

    # 在最终路径上已完成写入与着色
    
    print("\n" + "="*60)
    print("✅ Pipeline 执行完成！")
    print("="*60)
    print(f"\n📋 输出文件：")
    print(f"1. {output_path}")
    print(f"   - 带公式且着色版本")
    print(f"   - 基于 CONFIG.COLS_CONFIG 配置着色")
    print(f"\n💡 说明：仅输出带颜色版本，便于直接人工校验")


def load_json_from_llm_completion(response_str):
    response = response_str.strip()
    if response.startswith('```json'):
        response = response[7:]
    if response.startswith('```'):
        response = response[3:]
    if response.endswith('```'):
        response = response[:-3]
    response = response.strip()
    return json.loads(response)



def is_legal_date_batch(date_json: str) -> list:
    prompt = f"""
    当前日期：{datetime.now().strftime('%Y年%m月')}

    你是一个日期字符串合法性判断专家。  
    输入是一个 JSON 格式的 Excel 数据，包含人类手写的日期字符串和对应的行号。日期可能包含简单笔误但仍可解析为合法日期，或因错误被判定为非法日期。  
    你的任务是识别输入中所有非法日期的行号、原始日期文本和非法原因，并返回这些信息。

    **合法日期定义**：  
    可明确解析为有效日期（符合公历规则），即使包含简单笔误，例如：  
    - "2025年10月  28日" → 解析为 "2025年10月28日"  
    - "2025年10.28日" → 解析为 "2025年10月28日"  
    - "2025年1028" → 解析为 "2025年10月28日"  
    - "20251028" → 解析为 "2025年10月28日"  
    - "2025年10月5号" → 解析为 "2025年10月5日"  
    - "9.13" → 解析为 "2025年9月13日"  
    - "2025.10.1补" 解析为 "2025年10月1日"  

    **非法日期定义**：  
    1. 日期有歧义，无法明确解析，例如：  
       - "2025年9223" → 无法判断是9月22日还是9月23日  
    2. 日期明显无效（例如月份或日期超出有效范围），例如：  
       - "2025年13月28日" → 13月不存在  
       - "2025年2月30日" → 2月无30日  
    3. 不是当年数据，例如：  
       - "2024年10月28日" → 年份不对，当前时间是{int(datetime.now().strftime('%Y%m%d'))}年 
       - "2028年10月28日" → 年份不对，当前时间是 {int(datetime.now().strftime('%Y%m%d'))} 年

    **输入格式**：  
    JSON 数组，每项包含行号和日期字符串，例如：  
    ```json
    [
        {{ "行号": 44, "日期（必填）": "2025年9月21日" }},
        {{ "行号": 45, "日期（必填）": "2025年9117日" }}
    ]
    ```

    **任务， 现在开始回答我的问题，json格式返回**：  
    检查输入 `{date_json}` 中的每个日期字符串，找出所有非法日期。  
    - 如果存在非法日期，返回包含每项非法日期的行号、原始日期文本和原因的 JSON 数组。  
    - 如果没有非法日期，返回空数组 `[]`。  

    **返回格式**：  
    ```json
    [
        {{
            "行号": int,
            "日期（必填）": "原始非法日期文本",
            "reason": "判定为非法的原因"
        }}
    ]
    ```

    **示例**：  
    输入：  
    ```json
    [
        {{ "行号": 44, "日期（必填）": "2025年9月21日" }},
        {{ "行号": 45, "日期（必填）": "2025年9117日" }}
    ]
    ```  
    输出：  
    ```json
    [
        {{
            "行号": 45,
            "日期（必填）": "2025年9117日",
            "reason": "日期有歧义，无法解析为有效日期"
        }}
    ]
    ```  
    输入（全合法）：  
    ```json
    [
        {{ "行号": 44, "日期（必填）": "2025年9月21日" }}
    ]
    ```  
    输出：  
    ```json
    []
    ```
    """
    response = model_client.get_completion(prompt)
    try:
        response = parse_json_markdown(response)
        return response
    except:
        raise ValueError(prompt, response)

def check_date_for_df(df: pd.DataFrame):
    # 将日期列转换为字符串，避免 JSON 序列化问题
    df_copy = df[['行号', '日期（必填）']].copy()
    df_copy['日期（必填）'] = df_copy['日期（必填）'].astype(str)
    data = df_copy.to_dict(orient='records')
    data = json.dumps(data, ensure_ascii=False, indent=2)

    res = is_legal_date_batch(data)
    print('-----------------日期检测返回结果----------------')
    print(res)
    print('---------------------------------------------')
    if res:
        summary_text = '上传文件的日期[必填]字段有问题，请先解决日期问题再上传处理：\n'
        for item in res:
            row_name = item['行号']
            date_value = item['日期（必填）']
            reason = item['reason']
            summary_text += f"数据第{row_name}行日期有问题: {date_value} {reason}\n"
        return {
            'is_legal': False,
            'summary_text': summary_text
        }
    else:
        return {
            'is_legal': True,
            'summary_text': ''
        }
    
 

def process_ahead(df: pd.DataFrame, selected_halls: List[str]) -> Dict[str, Any]:
    """
    预校验函数，只校验数据量和日期格式

    Args:
        df: 输入的 DataFrame
        selected_halls: 选择的厅号列表

    Returns:
        Dict: 包含校验结果和错误信息的字典
    """
    # 1. 添加行号（如果没有的话）
    if '行号' not in df.columns:
        df['行号'] = df.index + 2

    # 2. 检查筛选后的数据是否为空
    if not selected_halls or len(selected_halls) == 0:
        return {
            'valid': False,
            'errors': ["必须至少选择一个厅号进行处理"]
        }

    # 拼接厅号筛选条件
    hall_filter = '|'.join(selected_halls)
    df_filtered_by_halls = df[df['厅号（必填）'].str.contains(hall_filter, na=False)]

    if len(df_filtered_by_halls) == 0:
        return {
            'valid': False,
            'errors': [f"筛选后的数据为空！Excel文件中未找到包含以下厅号的数据: {', '.join(selected_halls)}"]
        }

    # 3. 校验日期字段
    print('tolist-----------')
    print(df_filtered_by_halls['日期（必填）'].tolist())


    date_res = check_date_for_df(df_filtered_by_halls)

    if not date_res['is_legal']:
        # 将错误信息格式化为前端友好的格式
        error_lines = [line.strip() for line in date_res['summary_text'].strip().split('\n') if line.strip()]
        return {
            'valid': False,
            'errors': error_lines
        }

    return {
        'valid': True,
        'errors': []
    }


def get_date_str_from_text(text):
    PROMPT_DATE_EXTRACT = f'''
    当前时间是: {datetime.now().strftime('%Y年')}

    你是一个时间日期解析助手。请根据以下文本，返回这个文件合适的日期字符串。
    
    输入文本：{text}
    
    **重要规则：**
    1. 如果文件名中**没有明确的日期信息**，返回空字符串 ''
    2. 如果日期信息**有歧义**（如09223可能是0922或0923），返回空字符串 ''
    3. 如果日期信息**不完整**（如只有月份没有日期），返回空字符串 ''
    4. 只有**明确且无歧义**的日期才返回 YYYYMMDD 格式
    5. 支持多种日期格式的识别和转换
    
    返回格式为 YYYYMMDD 的日期字符串，不确定时返回空字符串 ''。
    
    示例：
    输入：映客10月03日打卡数据.xlsx， 且我们已知当年是2025年，则
    输出：20251003
    
    输入：主持打卡9.27.xlsx
    输出：20250927
    
    输入：打卡数据.xlsx
    输出：
    
    输入：主持打卡09223.xlsx
    这个有歧义
    输出：
    
    输入：10月打卡.xlsx
    输出：
    
    输入：主持打卡2024-10-03.xlsx
    输出：20241003
    
  

    **只返回日期字符串或空字符串，不要有任何其他文字！**
    '''
    try:
        response = model_client.get_completion(
            PROMPT_DATE_EXTRACT.format(text=text)
        )
        
        # 清理响应内容
        response = response.strip().replace("'", "")
        
        # 验证返回格式
        if response == '':
            return ''
        
        # 检查是否为8位数字格式
        if len(response) == 8 and response.isdigit():
            return response
        
        # 如果不符合格式，返回空字符串
        return ''
        
    except Exception as e:
        print(f"日期提取出错: {e}")
        return ''

 

 

 
if __name__ == '__main__':
 
 
    paths = list(Path('daily_data').glob('*.xlsx'))
    for p in paths:
        p = Path(p)
 
        output_path = p.with_suffix(f'.{CURRENT_MODEL}.v3.output.xlsx')
        # if 'v3' in p.stem:
        #     continue
        if 'output' in p.stem:
            continue 

        date_str_from_ui = get_date_str_from_text(p.stem)
        from global_config import STRICT_DATE_FILTER
        process_one_file(p, output_path=output_path, date_str_from_ui=date_str_from_ui, strict_date_filter=STRICT_DATE_FILTER)
        time.sleep(1)